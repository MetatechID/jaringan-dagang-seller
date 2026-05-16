#!/usr/bin/env python3
"""End-to-end smoke test for the catalog importer pipeline.

Generates synthetic XLSX fixtures matching each source adapter's column
convention, then runs parser → normalizer and asserts on the canonical
ImportedItem output. Does NOT touch the database; applier verification is
covered by Task #15 manual prod test.

Run:
    uv run python scripts/test-catalog-import.py
"""

from __future__ import annotations

import io
import sys
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook

from app.services.catalog_import.adapters import get_adapter, list_adapters
from app.services.catalog_import.normalizer import normalize
from app.services.catalog_import.parser import ParseError, parse_spreadsheet


GREEN = "\033[32m"
RED = "\033[31m"
YELLOW = "\033[33m"
RESET = "\033[0m"

failures: list[str] = []


def assert_eq(label: str, got, expected) -> None:
    if got == expected:
        print(f"  {GREEN}OK{RESET}  {label}: {got!r}")
    else:
        msg = f"  {RED}FAIL{RESET}  {label}: got {got!r}, expected {expected!r}"
        print(msg)
        failures.append(f"{label}: got {got!r} expected {expected!r}")


def _make_xlsx(headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_bigseller_fixture() -> None:
    """One product with 2 variants from BigSeller; price in 'Rp 12.500' format."""
    print(f"\n{YELLOW}== BigSeller ==", RESET)
    adapter = get_adapter("bigseller")
    bytes_ = _make_xlsx(
        ["Product ID", "Variation ID", "Product Name", "SKU", "Selling Price",
         "Stock", "Variation Type", "Variation Name", "Image URL",
         "Weight (g)", "Category", "Description"],
        [
            ["P1001", "V1", "Kerupuk Udang Premium", "KU-250", "Rp 12.500",
             50, "Size", "250g",
             "https://cf.bigseller.id/kerupuk-1.jpg,https://cf.bigseller.id/kerupuk-2.jpg",
             100, "Snacks", "Crispy prawn crackers"],
            ["P1001", "V2", "Kerupuk Udang Premium", "KU-500", "Rp 22.000",
             30, "Size", "500g", "https://cf.bigseller.id/kerupuk-1.jpg",
             200, "Snacks", "Crispy prawn crackers"],
        ],
    )
    headers, raw_rows, parse_warnings = parse_spreadsheet(bytes_, "test.xlsx")
    items, summary = normalize(raw_rows, dict(adapter.default_column_mapping), "bigseller")
    assert_eq("len(items)", len(items), 2)
    assert_eq("variant 1 price", items[0].price, Decimal(12500))
    assert_eq("variant 1 stock", items[0].stock, 50)
    assert_eq("variant 1 sku", items[0].sku_code, "KU-250")
    assert_eq("variant 1 parent_group_key", items[0].parent_group_key, "P1001")
    assert_eq("variant 2 parent_group_key", items[1].parent_group_key, "P1001")
    assert_eq("variant 1 images count", len(items[0].image_urls), 2)
    assert_eq("variant 1 weight g", items[0].weight_grams, 100)
    assert_eq("variant 1 errors", items[0].errors, [])
    assert_eq("summary", summary, {"new": 2, "update": 0, "warn": 0, "error": 0, "total": 2})


def test_shopee_fixture() -> None:
    print(f"\n{YELLOW}== Shopee ==", RESET)
    adapter = get_adapter("shopee")
    bytes_ = _make_xlsx(
        ["Parent SKU", "Variation SKU", "Product Name", "Price", "Stock",
         "Variation Name", "Option", "Cover Image", "Weight", "Category", "Product Description"],
        [
            ["SHP-001", "SHP-001-S", "Baju Kaos", "75000", 10, "Size", "S",
             "https://cf.shopee.co.id/x.jpg", 200, "Apparel", "Soft cotton tee"],
            ["SHP-001", "SHP-001-M", "Baju Kaos", "75000", 12, "Size", "M",
             "https://cf.shopee.co.id/x.jpg", 220, "Apparel", "Soft cotton tee"],
        ],
    )
    headers, raw_rows, _ = parse_spreadsheet(bytes_, "test.xlsx")
    items, summary = normalize(raw_rows, dict(adapter.default_column_mapping), "shopee")
    assert_eq("len(items)", len(items), 2)
    assert_eq("price", items[0].price, Decimal(75000))
    assert_eq("variant_value", items[0].variant_value, "S")
    assert_eq("parent_group_key", items[0].parent_group_key, "SHP-001")
    assert_eq("source_variant_id", items[0].source_variant_id, "SHP-001-S")
    assert_eq("errors", items[0].errors, [])


def test_tokopedia_indonesian_headers() -> None:
    print(f"\n{YELLOW}== Tokopedia (Bahasa headers) ==", RESET)
    adapter = get_adapter("tokopedia")
    bytes_ = _make_xlsx(
        ["Kode Produk Induk", "Kode Produk Varian", "Nama Produk", "SKU",
         "Harga", "Stok", "Tipe Varian", "Nama Varian", "URL Gambar",
         "Berat (g)", "Kategori", "Deskripsi"],
        [
            ["TKP-A", "TKP-A-1", "Sambal Roa", "SR-100", "Rp 35.000", 25,
             "Ukuran", "100ml", "https://tokopedia.cdn/r1.jpg", 150, "Saus", "Pedas khas Manado"],
        ],
    )
    headers, raw_rows, _ = parse_spreadsheet(bytes_, "test.xlsx")
    items, summary = normalize(raw_rows, dict(adapter.default_column_mapping), "tokopedia")
    assert_eq("len(items)", len(items), 1)
    assert_eq("name (Bahasa)", items[0].name, "Sambal Roa")
    assert_eq("price (Rp parse)", items[0].price, Decimal(35000))
    assert_eq("weight grams", items[0].weight_grams, 150)


def test_lazada_kg_to_grams() -> None:
    print(f"\n{YELLOW}== Lazada (kg → g conversion) ==", RESET)
    adapter = get_adapter("lazada")
    bytes_ = _make_xlsx(
        ["SellerSku", "VariationSku", "Name", "Price", "Quantity",
         "VariationName", "VariationValue", "Image1", "Package_weight",
         "PrimaryCategory", "Description"],
        [
            ["LZ-A", "LZ-A-1", "Coffee Bean", "85000", 50,
             "Roast", "Dark", "https://laz.cdn/1.jpg", "0.5", "Beverage", "Robusta from Toraja"],
        ],
    )
    headers, raw_rows, _ = parse_spreadsheet(bytes_, "test.xlsx")
    items, summary = normalize(raw_rows, dict(adapter.default_column_mapping), "lazada")
    assert_eq("Lazada 0.5 kg → 500 g", items[0].weight_grams, 500)


def test_generic_csv() -> None:
    print(f"\n{YELLOW}== Generic CSV ==", RESET)
    adapter = get_adapter("generic")
    csv = b"Parent ID,Variant ID,Name,SKU,Price,Stock,Variant Name,Variant Value,Image URL,Weight (g),Category,Description\n"
    csv += b"G1,V1,Roti Tawar,RT-1,15000,40,,,https://img.cdn/r.jpg,500,Bakery,\n"
    headers, raw_rows, _ = parse_spreadsheet(csv, "test.csv")
    items, summary = normalize(raw_rows, dict(adapter.default_column_mapping), "generic")
    assert_eq("len(items)", len(items), 1)
    assert_eq("price", items[0].price, Decimal(15000))


def test_errors_and_warnings() -> None:
    print(f"\n{YELLOW}== Errors + warnings detection ==", RESET)
    adapter = get_adapter("bigseller")
    bytes_ = _make_xlsx(
        ["Product ID", "Variation ID", "Product Name", "SKU", "Selling Price", "Stock"],
        [
            ["P1", "V1", "Good", "G-1", "12500", 10],
            ["P1", "V2", "", "G-1", "oops", ""],  # duplicate SKU + missing name + bad price + missing stock
            ["P2", "V3", "Free Item", "F-1", "0", 5],  # zero price → warning
        ],
    )
    headers, raw_rows, _ = parse_spreadsheet(bytes_, "test.xlsx")
    items, summary = normalize(raw_rows, dict(adapter.default_column_mapping), "bigseller")
    assert_eq("summary.error", summary["error"], 1)
    assert_eq("summary.warn (only row 3, row 2 is error so excluded)", summary["warn"], 1)
    assert_eq("summary.new (only row 1)", summary["new"], 1)
    assert_eq("row 3 warnings contain 'Price is 0'", "Price is 0" in items[2].warnings, True)
    assert_eq("row 2 errors contain duplicate SKU note",
              any("Duplicate SKU" in e for e in items[1].errors), True)


def test_multi_sheet_warning() -> None:
    print(f"\n{YELLOW}== Multi-sheet XLSX warning ==", RESET)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Products"
    ws1.append(["Name", "SKU", "Price", "Stock"])
    ws1.append(["X", "X1", "10000", 1])
    wb.create_sheet("Junk").append(["irrelevant"])
    buf = io.BytesIO()
    wb.save(buf)
    headers, rows, warnings = parse_spreadsheet(buf.getvalue(), "x.xlsx")
    assert_eq("has multi-sheet warning",
              any("sheets" in w for w in warnings), True)


def test_unknown_file_extension() -> None:
    print(f"\n{YELLOW}== Unknown extension → ParseError ==", RESET)
    try:
        parse_spreadsheet(b"x,y\n1,2", "file.txt")
        assert_eq("raised ParseError", False, True)
    except ParseError as e:
        assert_eq("raised ParseError", True, True)
        assert_eq("error mentions extension", "Unsupported" in str(e), True)


def test_csv_semicolon_dialect() -> None:
    print(f"\n{YELLOW}== CSV with semicolon delimiter (Indonesian Excel) ==", RESET)
    adapter = get_adapter("generic")
    csv = b"Parent ID;Variant ID;Name;SKU;Price;Stock\nG1;V1;Test;T1;10000;5"
    headers, rows, _ = parse_spreadsheet(csv, "test.csv")
    assert_eq("parsed headers count", len(headers), 6)
    items, _ = normalize(rows, dict(adapter.default_column_mapping), "generic")
    assert_eq("price parsed from semicolon CSV", items[0].price, Decimal(10000))


def test_all_adapters_register() -> None:
    print(f"\n{YELLOW}== All 5 adapters registered ==", RESET)
    names = sorted(a.name for a in list_adapters())
    assert_eq("adapter names", names, ["bigseller", "generic", "lazada", "shopee", "tokopedia"])


def main() -> int:
    test_all_adapters_register()
    test_bigseller_fixture()
    test_shopee_fixture()
    test_tokopedia_indonesian_headers()
    test_lazada_kg_to_grams()
    test_generic_csv()
    test_errors_and_warnings()
    test_multi_sheet_warning()
    test_unknown_file_extension()
    test_csv_semicolon_dialect()

    print()
    if failures:
        print(f"{RED}FAILED: {len(failures)} assertion(s){RESET}")
        for f in failures:
            print(f"  - {f}")
        return 1
    print(f"{GREEN}All assertions passed.{RESET}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
