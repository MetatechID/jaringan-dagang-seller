"""Spreadsheet parsing — XLSX/CSV bytes into a list of header-keyed dicts.

The parser is source-agnostic. It returns plain dicts; the adapter layer
applies source-specific column mappings and shape conventions.
"""

from __future__ import annotations

import csv
import io
from typing import Any

from openpyxl import load_workbook


class ParseError(Exception):
    """Raised when a file cannot be decoded or has no usable rows."""


def parse_spreadsheet(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    """Parse content into (headers, rows, warnings).

    Each row is a dict keyed by the header strings as-they-appear in row 1.
    Header cells are coerced to str and stripped of leading/trailing whitespace.
    Empty rows (all cells blank) are skipped.

    Returns warnings for non-fatal issues like multi-sheet XLSX files or
    encoding fallbacks.
    """

    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return _parse_xlsx(content)
    if lower.endswith(".csv"):
        return _parse_csv(content)
    raise ParseError(f"Unsupported file extension: {filename!r}. Use .xlsx or .csv")


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"Could not open XLSX: {e}") from e

    sheet_names = wb.sheetnames
    if len(sheet_names) > 1:
        warnings.append(
            f"File has {len(sheet_names)} sheets; only the first ('{sheet_names[0]}') is imported. "
            "Other sheets ignored."
        )

    ws = wb[sheet_names[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ParseError("File is empty")

    headers = [_clean_header(c) for c in header_row]
    rows: list[dict[str, Any]] = []
    for row_values in rows_iter:
        if all(_is_blank(v) for v in row_values):
            continue
        d = {}
        for h, v in zip(headers, row_values):
            if not h:
                continue
            d[h] = v
        rows.append(d)

    wb.close()
    if not rows:
        raise ParseError("File has no data rows after the header")
    return headers, rows, warnings


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    text: str | None = None
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            text = content.decode(encoding)
            if encoding != "utf-8":
                warnings.append(f"File decoded as {encoding}; UTF-8 is preferred.")
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ParseError("Could not decode file as UTF-8, latin-1, or cp1252")

    # csv.Sniffer can detect ; vs , — Indonesian Excel often exports semicolons.
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # fall back to comma

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    try:
        header_row = next(reader)
    except StopIteration:
        raise ParseError("File is empty")
    headers = [_clean_header(c) for c in header_row]

    rows: list[dict[str, Any]] = []
    for row_values in reader:
        if all(_is_blank(v) for v in row_values):
            continue
        d = {}
        for h, v in zip(headers, row_values):
            if not h:
                continue
            d[h] = v
        rows.append(d)

    if not rows:
        raise ParseError("File has no data rows after the header")
    return headers, rows, warnings


def _clean_header(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False
